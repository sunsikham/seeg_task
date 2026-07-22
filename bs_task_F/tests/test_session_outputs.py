import os
import tempfile
import unittest

from initiate import define_save

try:
    from openpyxl import load_workbook
    from save_func.save_results import (
        save_check_results_to_excel,
        save_results_to_excel,
    )
except ModuleNotFoundError:
    load_workbook = None
    save_check_results_to_excel = None
    save_results_to_excel = None


class SessionOutputTests(unittest.TestCase):
    def test_duplicate_subject_directory_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            save_dir = define_save(temp_dir, "P01")
            sentinel = os.path.join(save_dir, "keep.txt")
            with open(sentinel, "w", encoding="utf-8") as sentinel_file:
                sentinel_file.write("keep")

            with self.assertRaises(FileExistsError):
                define_save(temp_dir, "P01")
            with open(sentinel, encoding="utf-8") as sentinel_file:
                self.assertEqual(sentinel_file.read(), "keep")

    @unittest.skipIf(load_workbook is None, "openpyxl is not installed")
    def test_main_and_check_results_include_session_identifiers(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = save_results_to_excel(
                [{"trial_id": "T01"}],
                temp_dir,
                "main.xlsx",
                session_id="S01",
                neon_recording_id="R01",
            )
            check_path = save_check_results_to_excel(
                [{"phase": "phase1", "correct": True}],
                "food",
                temp_dir,
                "check.xlsx",
                session_id="S01",
                neon_recording_id="R01",
            )

            for path in (main_path, check_path):
                worksheet = load_workbook(path, read_only=True).active
                headers = [cell.value for cell in next(worksheet.iter_rows())]
                values = [cell.value for cell in next(worksheet.iter_rows())]
                row = dict(zip(headers, values))
                self.assertEqual(row["session_id"], "S01")
                self.assertEqual(row["neon_recording_id"], "R01")


if __name__ == "__main__":
    unittest.main()
