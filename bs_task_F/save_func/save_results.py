import os
from openpyxl import Workbook


from openpyxl import Workbook
import os

from openpyxl import Workbook
import os

def _save_workbook_atomic(wb, file_path):
    temp_file_path = f"{file_path}.tmp"

    try:
        wb.save(temp_file_path)
        os.replace(temp_file_path, file_path)
    finally:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)


def save_results_to_excel(results, save_dir="data", filename="results.xlsx"):

    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    # =========================
    # 헤더 (새로운 자료구조에 맞게 변경)
    # =========================
    headers = [
        "index",
        "trial_id",       # JSON에 적어둔 문제 번호
        "domain",         # 화면 출력용 대분류 (food, gene, habitat)
        "task_type",      # 세부 유형 (Food, Tree, Conflict, Habitat, Tree-conflict 등)
        "question_target",
        "question_text",

        "premise",
        "option1",
        "option2",
        
        "rationale",      # 정답의 근거 (거리 비교 텍스트)
        "checkpoint1",    # Food 태스크용 포인트
        "checkpoint2",    # Food 태스크용 포인트

        "correct_answer",
        "response",
        "rt1",
        "rt2",
        "question_onset_s",
        "premise_onset_s",
        "option_left_onset_s",
        "option_right_onset_s",
        "choice_onset_s",
        "response_detected_s",
        "is_correct"
    ]
    ws.append(headers)

    # =========================
    # 데이터
    # =========================
    idx = 0
    for r in results:
        idx+=1
        ws.append([
            idx,
            r.get("trial_id", ""),
            r.get("domain", ""),
            r.get("task_type", ""),
            r.get("question_target", ""),
            r.get("question_text", ""),

            r.get("premise", ""),
            r.get("option1", ""),
            r.get("option2", ""),
            
            r.get("rationale", ""),    # 값이 없는 태스크는 빈칸으로 들어갑니다
            r.get("checkpoint1", ""),  # 값이 없는 태스크는 빈칸으로 들어갑니다
            r.get("checkpoint2", ""),  # 값이 없는 태스크는 빈칸으로 들어갑니다

            r.get("correct", ""),      # JSON의 correct 키 매핑
            r.get("response", ""),
            r.get("rt1", ""),
            r.get("rt2", ""),
            r.get("question_onset_s", ""),
            r.get("premise_onset_s", ""),
            r.get("option_left_onset_s", ""),
            r.get("option_right_onset_s", ""),
            r.get("choice_onset_s", ""),
            r.get("response_detected_s", ""),
            r.get("is_correct", "")
        ])

    _save_workbook_atomic(wb, file_path)

    return file_path


def save_check_results_to_excel(
    results,
    task_type,
    save_dir="data",
    filename="check_results.xlsx",
):
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "check_results"

    headers = [
        "attempt_index",
        "task_type",
        "phase",
        "animal",
        "slot_idx",
        "slot_label",
        "correct",
        "rt_sec",
        "rt_frames",
    ]
    ws.append(headers)

    for attempt_index, result in enumerate(results, start=1):
        ws.append([
            attempt_index,
            task_type,
            result.get("phase", ""),
            result.get("animal", ""),
            result.get("slot_idx", ""),
            result.get("slot_label", ""),
            result.get("correct", ""),
            result.get("rt_sec", ""),
            result.get("rt_frames", ""),
        ])

    _save_workbook_atomic(wb, file_path)

    return file_path

def save_results_to_excel_A(results, save_dir="data", filename="results.xlsx"):

    # =========================
    # 1. 폴더 생성
    # =========================
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, filename)

    # =========================
    # 2. 워크북 생성
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    # =========================
    # 3. 헤더
    # =========================
    headers = [
        "trial_index",
        "animal_a",
        "animal_b",
        "correct_answer",   # 1 (O) / 0 (X)
        "response",         # 1 / 0 / None
        "rt",
        "is_correct"
    ]
    ws.append(headers)

    # =========================
    # 4. 데이터 기록
    # =========================
    for i, r in enumerate(results, start=1):

        ws.append([
            i,
            r.get("animal_a"),
            r.get("animal_b"),
            r.get("correct_answer"),
            r.get("response"),
            r.get("rt"),
            r.get("is_correct")
        ])

    # =========================
    # 5. 저장
    # =========================
    wb.save(file_path)

    print(f"Saved to: {file_path}")
    return file_path
